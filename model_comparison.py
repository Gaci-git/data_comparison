import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import os

# Read CSV files -- Copy paste your file paths for csv you'll be comparing

v1_path = r"your_csv_file"
v1 = pd.read_csv(v1_path, sep=';', encoding='UTF-8', on_bad_lines='skip')

v2_path = r"your_csv_file"
v2 = pd.read_csv(v2_path, sep=';', encoding='UTF-8', on_bad_lines='skip')

# Make all column names lowercase for comparison
v1.columns = v1.columns.str.lower()
v2.columns = v2.columns.str.lower()

# Remove technical columns
cols_to_remove = ['src_popul_ts', 'popul_ts', 'popul_ts_last']
v1 = v1.drop(columns=[col for col in cols_to_remove if col in v1.columns])
v2 = v2.drop(columns=[col for col in cols_to_remove if col in v2.columns])

# Read column names and make a list from them.
common_cols = [col for col in v1.columns if col in v2.columns]

# Sort you dataframe by the common_cols list
sort_columns = common_cols
v1 = v1.sort_values(by=sort_columns).reset_index(drop=True)
v2 = v2.sort_values(by=sort_columns).reset_index(drop=True)

# Here you can change your sample size. Limit the number of rows for comparison (for performance)
v1 = v1.head(1000)
v2 = v2.head(1000)

# Find rows that missmatch - for example they are only in v1 or v2
diff_v1 = pd.merge(v1, v2, how='outer', indicator=True).query('_merge == "left_only"').drop('_merge', axis=1)
diff_v2 = pd.merge(v1, v2, how='outer', indicator=True).query('_merge == "right_only"').drop('_merge', axis=1)

# Add index for comparison - will be used as key for merging v1 and v2 dataframes
v1['row_id'] = v1.index
v2['row_id'] = v2.index
merged = pd.merge(v1, v2, on='row_id', how='outer', suffixes=('_v1', '_v2'))

# Finding columns with values that differ (in order to add that additional sheet with different columns) 
def find_differences(row):
    diffs = []
    for col in common_cols:
        val1 = row.get(f"{col}_v1")
        val2 = row.get(f"{col}_v2")
        if pd.notna(val1) and pd.notna(val2) and str(val1) != str(val2):
            diffs.append(col)
    return ', '.join(diffs) if diffs else None

merged['diff_columns'] = merged.apply(find_differences, axis=1)
rows_with_diffs = merged[merged['diff_columns'].notna()]

# Writing our findings into excel report
wb = Workbook()

# Summary sheet
summary_sheet = wb.active
summary_sheet.title = "Summary"
summary_sheet.append(["Comparison", "Count"])
summary_sheet.append(["Rows only in version 1", len(diff_v1)])
summary_sheet.append(["Rows only in version 2", len(diff_v2)])
summary_sheet.append(["Rows with column-level differences", len(rows_with_diffs)])
for cell in summary_sheet[1]:
    cell.font = Font(bold=True)

# function to write other 3 remaining sheets (v1 data, v2 data, different column) and format the excel a bit. 
def write_diff_sheet(sheet_name, data):
    sheet = wb.create_sheet(title=sheet_name)
    if not data.empty:
        for c_idx, column_name in enumerate(data.columns, start=1):
            header_cell = sheet.cell(row=1, column=c_idx, value=column_name)
            header_cell.font = Font(bold=True)
            header_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        for r_idx, row in enumerate(data.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

# Adding sheets for differences found in each version and column differences
write_diff_sheet("Only_in_v1", diff_v1)
write_diff_sheet("Only_in_v2", diff_v2)
write_diff_sheet("Column_Differences", rows_with_diffs[['row_id', 'diff_columns']])

# Naming accordingly and saving the excel report.
csv_filename = os.path.splitext(os.path.basename(v1_path))[0]
excel_report_name = f"Differences_Report_{csv_filename}.xlsx"
wb.save(excel_report_name)
print(f"Differences exported to '{excel_report_name}'")
